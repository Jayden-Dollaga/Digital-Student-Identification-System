###############################################################################
#  gui/attendance_report_page.py
#  AS608 Fingerprint Attendance System
#
#  "Reports" tab — a searchable, date-filtered attendance report for
#  students, one row per (student, day), with Time-In / Time-Out.
#  Mirrors the layout of the original web BFEAS "Attendance Report" screen,
#  built for this app's CustomTkinter shell using real database data.
###############################################################################

import sys
from pathlib import Path

PYTHON_ROOT = Path(__file__).resolve().parents[1]
if str(PYTHON_ROOT) not in sys.path:
    sys.path.insert(0, str(PYTHON_ROOT))

from datetime import datetime, timedelta
from tkinter import ttk, messagebox

import customtkinter as ctk

from core.database import get_daily_attendance_summary
from gui.theme import get_theme_colors

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:  # pragma: no cover
    openpyxl = None

from core.utils import get_export_path, timestamp_filename


REPORT_HEADERS = ["No.", "Student Name", "Student No.", "Grade", "Section",
                  "Date", "Day", "Time-In", "Time-Out", "Status"]

HEADER_BLUE = "#3f7fbf"
SIDEBAR_BLUE = "#123a5e"
SIDEBAR_BLUE_LIGHT = "#1c4a72"
SIDEBAR_SECTION = "#0c2c47"
ACCENT_BLUE = "#2f7bc4"
BODY_BG = "#eef1f5"
CARD_BG = "#ffffff"
TEXT_DARK = "#2b2b2b"
BORDER = "#dfe3e8"


class AttendanceReportPage:
    """Reports tab: date-range attendance summary, one row per student/day."""

    def __init__(self, app):
        self.app = app
        self.container = None
        self.tree = None
        self.from_entry = None
        self.to_entry = None
        self.search_entry = None
        self.show_var = None
        self.showing_label = None
        self.page_label = None

        self._all_rows = []      # raw rows from the DB for the current date range
        self._filtered_rows = []  # after search filter
        self.current_page = 1

    # ------------------------------------------------------------------
    # Build
    # ------------------------------------------------------------------
    def build(self, parent):
        self.container = ctk.CTkFrame(parent, fg_color="transparent")
        self.container.grid(sticky="nsew")
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)
        self.container.grid_rowconfigure(2, weight=1)

        header = ctk.CTkFrame(self.container, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 10))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="Attendance Report", text_color=TEXT_DARK,
                     font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(
            header, text="⬇ Export to Excel", width=150, fg_color=ACCENT_BLUE,
            hover_color="#265f96", command=self.export_to_excel,
        ).grid(row=0, column=1, sticky="e")

        filter_card = ctk.CTkFrame(
            self.container, corner_radius=8, border_width=1, border_color=BORDER,
            fg_color=CARD_BG,
        )
        filter_card.grid(row=1, column=0, sticky="ew", padx=4, pady=(0, 10))

        today = datetime.now().strftime("%Y-%m-%d")
        month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

        ctk.CTkLabel(filter_card, text="From:", text_color=TEXT_DARK,
                     font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, padx=(14, 6), pady=14, sticky="w"
        )
        self.from_entry = ctk.CTkEntry(filter_card, placeholder_text="yyyy-mm-dd", width=140)
        self.from_entry.insert(0, month_ago)
        self.from_entry.grid(row=0, column=1, padx=(0, 16), pady=14)

        ctk.CTkLabel(filter_card, text="To:", text_color=TEXT_DARK,
                     font=("Segoe UI", 12, "bold")).grid(
            row=0, column=2, padx=(0, 6), pady=14, sticky="w"
        )
        self.to_entry = ctk.CTkEntry(filter_card, placeholder_text="yyyy-mm-dd", width=140)
        self.to_entry.insert(0, today)
        self.to_entry.grid(row=0, column=3, padx=(0, 16), pady=14)

        ctk.CTkButton(
            filter_card, text="View Reports", width=130, fg_color=ACCENT_BLUE,
            hover_color="#265f96", command=self.refresh,
        ).grid(row=0, column=4, padx=(0, 14), pady=14)

        table_card = ctk.CTkFrame(
            self.container, corner_radius=8, border_width=1, border_color=BORDER,
            fg_color=CARD_BG,
        )
        table_card.grid(row=2, column=0, sticky="nsew", padx=4, pady=(0, 4))
        table_card.grid_columnconfigure(0, weight=1)
        table_card.grid_rowconfigure(2, weight=1)

        controls = ctk.CTkFrame(table_card, fg_color="transparent")
        controls.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        controls.grid_columnconfigure(2, weight=1)

        ctk.CTkLabel(controls, text="Show", text_color=TEXT_DARK).grid(row=0, column=0, sticky="w")
        self.show_var = ctk.StringVar(value="10")
        show_menu = ctk.CTkOptionMenu(
            controls, values=["10", "25", "50", "100"], variable=self.show_var,
            width=70, command=lambda _=None: self._render_page(reset_page=True),
        )
        show_menu.grid(row=0, column=1, padx=8)
        ctk.CTkLabel(controls, text="entries", text_color=TEXT_DARK).grid(
            row=0, column=1, padx=(70, 0), sticky="w"
        )

        search_frame = ctk.CTkFrame(controls, fg_color="transparent")
        search_frame.grid(row=0, column=3, sticky="e")
        ctk.CTkLabel(search_frame, text="Search:", text_color=TEXT_DARK).pack(side="left", padx=(0, 6))
        self.search_entry = ctk.CTkEntry(search_frame, width=200)
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", lambda e: self._apply_search())

        style_name = "AttendanceReport.Treeview"
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            style_name, background="white", fieldbackground="white",
            rowheight=30, font=("Segoe UI", 10), borderwidth=0,
        )
        style.configure(
            f"{style_name}.Heading", font=("Segoe UI", 10, "bold"),
            background="#f4f6f8", foreground=TEXT_DARK, relief="flat",
        )
        style.map(style_name, background=[("selected", "#dbeafe")])

        tree_frame = ctk.CTkFrame(table_card, fg_color="transparent")
        tree_frame.grid(row=2, column=0, sticky="nsew", padx=14, pady=(0, 6))
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        columns = ("no", "name", "student_no", "grade", "section",
                   "date", "day", "time_in", "time_out", "status")
        widths = [40, 180, 100, 70, 100, 100, 90, 90, 90, 90]

        self.tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", style=style_name,
        )
        for col, head, width in zip(columns, REPORT_HEADERS, widths):
            self.tree.heading(col, text=head)
            self.tree.column(col, width=width, anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        footer = ctk.CTkFrame(table_card, fg_color="transparent")
        footer.grid(row=3, column=0, sticky="ew", padx=14, pady=(0, 14))
        footer.grid_columnconfigure(0, weight=1)

        self.showing_label = ctk.CTkLabel(
            footer, text="", text_color="#5a6672", font=("Segoe UI", 11)
        )
        self.showing_label.grid(row=0, column=0, sticky="w")

        pagination = ctk.CTkFrame(footer, fg_color="transparent")
        pagination.grid(row=0, column=1, sticky="e")
        ctk.CTkButton(pagination, text="Previous", width=80, fg_color="transparent",
                      border_width=1, command=self._prev_page).pack(side="left", padx=2)
        self.page_label = ctk.CTkLabel(pagination, text="1", width=30)
        self.page_label.pack(side="left", padx=6)
        ctk.CTkButton(pagination, text="Next", width=60, fg_color="transparent",
                      border_width=1, command=self._next_page).pack(side="left", padx=2)

        self.refresh()
        return self.container

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------
    def refresh(self):
        """Re-query the database for the currently selected date range."""
        if self.tree is None or not self.tree.winfo_exists():
            return

        from_date = self.from_entry.get().strip()
        to_date = self.to_entry.get().strip()

        if not self._valid_date(from_date) or not self._valid_date(to_date):
            messagebox.showerror(
                "Invalid Date",
                "Please enter dates in yyyy-mm-dd format for both From and To.",
                parent=self.app,
            )
            return

        try:
            self._all_rows = get_daily_attendance_summary(from_date, to_date)
        except Exception as exc:
            messagebox.showerror("Report Error", f"Could not load attendance report: {exc}", parent=self.app)
            self._all_rows = []

        self.current_page = 1
        self._apply_search()

    @staticmethod
    def _valid_date(value):
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def _apply_search(self):
        query = self.search_entry.get().strip().lower() if self.search_entry else ""
        if not query:
            self._filtered_rows = list(self._all_rows)
        else:
            self._filtered_rows = [
                row for row in self._all_rows
                if query in " ".join(str(v).lower() for v in row.values())
            ]
        self._render_page(reset_page=True)

    # ------------------------------------------------------------------
    # Rendering / pagination
    # ------------------------------------------------------------------
    def _render_page(self, reset_page=False):
        if reset_page:
            self.current_page = 1

        try:
            page_size = int(self.show_var.get())
        except Exception:
            page_size = 10

        total = len(self._filtered_rows)
        total_pages = max(1, (total + page_size - 1) // page_size)
        self.current_page = min(max(1, self.current_page), total_pages)

        start = (self.current_page - 1) * page_size
        end = start + page_size
        page_rows = self._filtered_rows[start:end]

        self.tree.delete(*self.tree.get_children())
        for idx, row in enumerate(page_rows, start=start + 1):
            date_str = row.get("date", "")
            try:
                day_name = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")
            except Exception:
                day_name = "N/A"

            status = "Present" if row.get("scan_count", 0) >= 1 else "N/A"
            time_out = row.get("time_out") or "—"

            self.tree.insert("", "end", values=(
                idx,
                row.get("student_name", "N/A"),
                row.get("student_no", "N/A"),
                row.get("grade", "N/A"),
                row.get("section", "N/A"),
                date_str,
                day_name,
                row.get("time_in", "N/A"),
                time_out,
                status,
            ))

        shown = len(page_rows)
        self.showing_label.configure(
            text=f"Showing {0 if shown == 0 else start + 1} to {start + shown} of {total} entries"
        )
        self.page_label.configure(text=f"{self.current_page} / {total_pages}")

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._render_page()

    def _next_page(self):
        try:
            page_size = int(self.show_var.get())
        except Exception:
            page_size = 10
        total_pages = max(1, (len(self._filtered_rows) + page_size - 1) // page_size)
        if self.current_page < total_pages:
            self.current_page += 1
            self._render_page()

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def export_to_excel(self):
        if openpyxl is None:
            messagebox.showwarning(
                "Excel Export Unavailable",
                "openpyxl is not installed. Run: pip install openpyxl",
                parent=self.app,
            )
            return

        if not self._filtered_rows:
            messagebox.showinfo("No Data", "There is nothing to export for this date range.", parent=self.app)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.append(REPORT_HEADERS)

            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="1F3864")
                cell.alignment = Alignment(horizontal="center")

            for idx, row in enumerate(self._filtered_rows, start=1):
                date_str = row.get("date", "")
                try:
                    day_name = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")
                except Exception:
                    day_name = "N/A"
                status = "Present" if row.get("scan_count", 0) >= 1 else "N/A"
                ws.append([
                    idx,
                    row.get("student_name", "N/A"),
                    row.get("student_no", "N/A"),
                    row.get("grade", "N/A"),
                    row.get("section", "N/A"),
                    date_str,
                    day_name,
                    row.get("time_in", "N/A"),
                    row.get("time_out") or "—",
                    status,
                ])

            for col_cells in ws.columns:
                ws.column_dimensions[col_cells[0].column_letter].width = 16

            filename = timestamp_filename("attendance_report", "xlsx")
            filepath = get_export_path(filename)
            wb.save(filepath)

            self.app.log_message(f"Attendance report exported to {filepath}")
            messagebox.showinfo("Export Successful", f"Report saved to:\n{filepath}", parent=self.app)
        except Exception as exc:
            messagebox.showerror("Export Error", f"Could not export report: {exc}", parent=self.app)


def build_reports_tab(app, parent):
    """Entry point mirroring build_statistics_tab / build_log_tab conventions."""
    page = getattr(app, "attendance_report_page", None)
    if page is None:
        page = AttendanceReportPage(app)
        app.attendance_report_page = page
    return page.build(parent)


class BFEASApp2(ctk.CTk):
    """Standalone window wrapper for the ported BFEAS reports UI."""

    def __init__(self):
        super().__init__()
        self.title("Biometric Fingerprint Employee Attendance System")
        self.geometry("1360x740")
        self.minsize(1100, 700)
        self.configure(fg_color=BODY_BG)

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self._build_topbar()
        self._build_sidebar()
        self._build_main_content()

        self.attendance_report_page = AttendanceReportPage(self)
        self.attendance_report_page.build(self.content_container)

    def _build_topbar(self):
        topbar = ctk.CTkFrame(self, fg_color=HEADER_BLUE, height=56, corner_radius=0)
        topbar.grid(row=0, column=0, columnspan=2, sticky="ew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(topbar, text="☰", text_color="white",
                     font=("Segoe UI", 18)).grid(row=0, column=0, padx=(18, 6), pady=10)
        ctk.CTkLabel(topbar, text="Biometric Fingerprint Employee Attendance System",
                     text_color="white", font=("Segoe UI", 15, "bold")).grid(
            row=0, column=1, sticky="w", pady=10
        )

        profile_frame = ctk.CTkFrame(topbar, fg_color="transparent")
        profile_frame.grid(row=0, column=2, sticky="e", padx=18)
        ctk.CTkLabel(profile_frame, text="👤", text_color=HEADER_BLUE, fg_color="white",
                     width=28, height=28, corner_radius=14,
                     font=("Segoe UI", 14)).grid(row=0, column=0, padx=(0, 8))
        ctk.CTkLabel(profile_frame, text="Cletus Igbe", text_color="white",
                     font=("Segoe UI", 13)).grid(row=0, column=1)

    def _build_sidebar(self):
        sidebar = ctk.CTkFrame(self, fg_color=SIDEBAR_BLUE, width=230, corner_radius=0)
        sidebar.grid(row=1, column=0, sticky="ns")
        sidebar.grid_propagate(False)

        profile = ctk.CTkFrame(sidebar, fg_color=SIDEBAR_BLUE, height=90)
        profile.pack(fill="x", pady=(18, 6))
        ctk.CTkLabel(profile, text="👤", fg_color="#cfd8e3", text_color=SIDEBAR_BLUE,
                     width=54, height=54, corner_radius=27,
                     font=("Segoe UI", 24)).pack()
        ctk.CTkLabel(profile, text="Cletus Igbe", text_color="white",
                     font=("Segoe UI", 14, "bold")).pack(pady=(6, 0))
        status = ctk.CTkFrame(profile, fg_color=SIDEBAR_BLUE)
        status.pack()
        ctk.CTkLabel(status, text="●", text_color="#3ddc63",
                     font=("Segoe UI", 10)).pack(side="left")
        ctk.CTkLabel(status, text=" Online", text_color="#cfe0f0",
                     font=("Segoe UI", 11)).pack(side="left")

        ctk.CTkFrame(sidebar, fg_color=SIDEBAR_BLUE_LIGHT, height=1).pack(fill="x", pady=8)

        self._nav_button(sidebar, "🏠 Dashboard", active=False)
        self._section_label(sidebar, "MANAGE")
        self._nav_button(sidebar, "👤 Employee", active=False)
        self._section_label(sidebar, "REPORTS")
        self._nav_button(sidebar, "👥 Attendance List", active=False)
        self._nav_button(sidebar, "📄 Attendance Report", active=True)

    def _section_label(self, parent, text):
        ctk.CTkLabel(parent, text=text, text_color="#7fa0bd",
                     font=("Segoe UI", 11, "bold"), anchor="w").pack(fill="x", padx=18, pady=(16, 4))

    def _nav_button(self, parent, text, active=False):
        fg = ACCENT_BLUE if active else SIDEBAR_BLUE
        hover = ACCENT_BLUE if active else SIDEBAR_BLUE_LIGHT
        ctk.CTkButton(parent, text=text, anchor="w", fg_color=fg, hover_color=hover,
                      text_color="white", corner_radius=0, height=42,
                      font=("Segoe UI", 13), command=lambda: None).pack(fill="x")

    def _build_main_content(self):
        main = ctk.CTkFrame(self, fg_color=BODY_BG, corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew")
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(main, fg_color=BODY_BG)
        header.grid(row=0, column=0, sticky="ew", padx=30, pady=(22, 10))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="Reports", text_color=TEXT_DARK,
                     font=("Segoe UI", 26, "bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(header, text="🏠 Home  >  Reports", text_color="#7a8794",
                     font=("Segoe UI", 12)).grid(row=0, column=1, sticky="e")

        self.content_container = ctk.CTkFrame(main, fg_color="transparent")
        self.content_container.grid(row=1, column=0, sticky="nsew", padx=30, pady=(0, 20))
        self.content_container.grid_columnconfigure(0, weight=1)
        self.content_container.grid_rowconfigure(0, weight=1)

    def log_message(self, message):
        print(message)


def main():
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")

    app = BFEASApp2()
    app.mainloop()


if __name__ == "__main__":
    main()